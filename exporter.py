from io import BytesIO
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from config import SHIFT_COLORS


def export_workbook(schedule_df, manpower_df, person_df, issues_df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        schedule_df.to_excel(writer, sheet_name="班表")
        manpower_df.to_excel(writer, sheet_name="每日人力", index=False)
        person_df.to_excel(writer, sheet_name="個人統計", index=False)
        issues_df.to_excel(writer, sheet_name="違規檢查", index=False)

        wb = writer.book
        for ws in wb.worksheets:
            style_sheet(ws)
        color_schedule_sheet(wb["班表"])
    output.seek(0)
    return output.getvalue()


def style_sheet(ws):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.freeze_panes = "B2"
    for col in ws.columns:
        max_len = 8
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value) + 2, 24))
        ws.column_dimensions[col_letter].width = max_len


def color_schedule_sheet(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            value = str(cell.value) if cell.value is not None else ""
            if value in SHIFT_COLORS:
                cell.fill = PatternFill("solid", fgColor=SHIFT_COLORS[value])
                if value == "N":
                    cell.font = Font(bold=True)
